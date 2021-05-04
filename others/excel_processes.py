import os 
from openpyxl import Workbook,load_workbook

os.chdir("..")

wb = load_workbook("statistics.xlsx")
ws = wb.active

def excel_clear():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    for row in ws:
        for cell in row:
            cell.value = None
    ws['B1'] = '1. çözüm ortalama işlenme süresi'
    ws['C1'] = '2. çözüm ortalama işlenme süresi'
    ws['D1'] = '1. çözüm ilk mesaj işlenme süresi'
    ws['E1'] = '2. çözüm ilk mesaj işlenme süresi'
    ws['F1'] = '1. çözüm için en yüksek ilk bekleme süresi kullanıcısı'
    ws['G1'] = '1. çözüm için en yüksek ilk bekleme süresi'
    ws['H1'] = '2. çözüm için en yüksek ilk bekleme süresi kullanıcısı'
    ws['I1'] = '2. çözüm için en yüksek ilk bekleme süresi'
    wb.save("statistics.xlsx")

def excel_read():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    for row in ws:
        for cell in row:
            print(cell.value, " ", end='')
        print()

while True:
    inpt = input()
    if inpt == 'c':
        excel_clear()
    if inpt == 'r':
        excel_read()