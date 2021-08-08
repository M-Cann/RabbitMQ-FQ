import os 
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

#os.chdir("..")

queues_names=['A','B','C','D','E','F','G','H','I','M']
other_value=['Toplam gelen mesaj sayısı', 'Toplam geçen zaman', 'Ortalaması', 'İlk mesaj için en çok bekleyen kuyruk', 'Beklediği süre']
wb = Workbook()
wb.save("statistics.xlsx")
wb2 = Workbook()
wb2.save("statistics2.xlsx")

def scenarios_count():
    path = "scenarios"
    count = os.listdir(path)
    return(len(count))


def excel_çözümsüz_clear():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    wb2 = load_workbook("statistics2.xlsx")
    ws2 = wb2.active
    sc_count = scenarios_count()
    for i in range(2,2+(sc_count*6), 6):
        column1=get_column_letter(i)
        column2=get_column_letter(i+1)
        for j in range (4,13):
            cell1 = column1+str(j)
            cell2 = column2+str(j)
            ws[cell1]=None
            ws[cell2]=None
    for i in range(2,2+(sc_count*3), 3):
        column=get_column_letter(i)
        for j in range (3,7):
            cell = column+str(j)
            ws2[cell]=None

def excel_block_clear():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    wb2 = load_workbook("statistics2.xlsx")
    ws2 = wb2.active
    sc_count = scenarios_count()
    for i in range(4,4+(sc_count*6), 6):
        column1=get_column_letter(i)
        column2=get_column_letter(i+1)
        for j in range (4,13):
            cell1 = column1+str(j)
            cell2 = column2+str(j)
            ws[cell1]=None
            ws[cell2]=None
    for i in range(3,3+(sc_count*3), 3):
        column=get_column_letter(i)
        for j in range (3,7):
            cell = column+str(j)
            ws2[cell]=None

def excel_select_clear():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    wb2 = load_workbook("statistics2.xlsx")
    ws2 = wb2.active
    sc_count = scenarios_count()
    for i in range(6,6+(sc_count*6), 6):
        column1=get_column_letter(i)
        column2=get_column_letter(i+1)
        for j in range (4,13):
            cell1 = column1+str(j)
            cell2 = column2+str(j)
            ws[cell1]=None
            ws[cell2]=None
    for i in range(3,3+(sc_count*3), 3):
        column=get_column_letter(i)
        for j in range (3,7):
            cell = column+str(j)
            ws2[cell]=None


def excel_clear():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    for row in ws:
        for cell in row:
            cell.value = None
    sc_count = scenarios_count()
    j=0
    for i in range(2,2+(sc_count*6), 6):
        ws.merge_cells(start_column=i, end_column=i+5, start_row=1, end_row=1)
        cell=str(get_column_letter(i))+str(1)
        ws[cell] = 'Seneryo ' + str(j)
        j=j+1
    for j in range(2,2+(sc_count*6), 6):
        ws.merge_cells(start_column=j, end_column=j+1, start_row=2, end_row=2)
        ws.merge_cells(start_column=j+2, end_column=j+3, start_row=2, end_row=2)
        ws.merge_cells(start_column=j+4, end_column=j+5, start_row=2, end_row=2)
        cell1=str(get_column_letter(j))+str(2)
        cell2=str(get_column_letter(j+2))+str(2)
        cell3=str(get_column_letter(j+4))+str(2)
        ws[cell1] = 'Çözümsüz'
        ws[cell2] = 'Block Connect'
        ws[cell3] = 'Select Connect'
    for k in range(2,2+(sc_count*6), 2):
        cell1=str(get_column_letter(k))+str(3)
        cell2=str(get_column_letter(k+1))+str(3)
        ws[cell1] = 'İlk mesajın süresi'
        ws[cell2] = 'Ortalama mesaj işleme süresi'
    for l in range(0,len(queues_names)):
        cell='A'+str(l+4)
        ws[cell]=queues_names[l]
    wb.save("statistics.xlsx")

    wb2 = load_workbook("statistics2.xlsx")
    ws2 = wb2.active
    for row in ws2:
        for cell in row:
            cell.value = None
    sc_count = scenarios_count()
    j=0
    for i in range(2,2+(sc_count*3), 3):
        ws2.merge_cells(start_column=i, end_column=i+2, start_row=1, end_row=1)
        cell=str(get_column_letter(i))+str(1)
        ws2[cell] = 'Seneryo ' + str(j)
        j=j+1
    for j in range(2,2+(sc_count*3), 3):
        cell1=str(get_column_letter(j))+str(2)
        cell2=str(get_column_letter(j+1))+str(2)
        cell3=str(get_column_letter(j+2))+str(2)
        ws2[cell1] = 'Çözümsüz'
        ws2[cell2] = 'Block Connect'
        ws2[cell3] = 'Select Connect'
    for l in range(0,len(other_value)):
        cell='A'+str(l+3)
        ws2[cell]=other_value[l]
    wb2.save("statistics2.xlsx")


def excel_read():
    wb = load_workbook("statistics.xlsx")
    ws = wb.active
    for row in ws:
        for cell in row:
            print(cell.value, " ", end='')
        print()
    wb2 = load_workbook("statistics2.xlsx")
    ws2 = wb2.active
    for row in ws2:
        for cell in row:
            print(cell.value, " ", end='')
        print()

while True:
    inpt = input()
    if inpt == 'c':
        excel_clear()
    if inpt == 'ç':
        excel_çözümsüz_clear()
    if inpt == 'b':
        excel_block_clear()
    if inpt == 's':
        excel_select_clear()
    if inpt == 'r':
        excel_read()