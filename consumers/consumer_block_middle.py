import os 
import pika
import time
import json  
import threading   
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

connection = pika.BlockingConnection(pika.ConnectionParameters(host='localhost'))
channel = connection.channel()
channel.queue_declare(queue='fair_queue', durable=True)
channel.basic_qos(prefetch_count=1)

#os.chdir("..")
first_message = {}
statistics_dic = {}
total_time = {}
queue_start_time = {}
queue_finish_time = {}
avarage=[] #Bu kısma bak fazladan hesap yapılıyor daha çok istatistik çıkarılabilir
queues = []

def basic_consume():
    def callback(ch, method, properties, body):
        receive_time = time.time()  
        received_data = json.loads(body)
        user = received_data['User']
        start_time = received_data['Time']
        process_time=receive_time-start_time
        if statistics_dic.get(user)==None:
            statistics_dic[user]=1
        else:
            statistics_dic[user]=statistics_dic.get(user)+1
        if first_message.get(user)==None:
            first_message[user]=process_time
        if total_time.get(user)==None:
            total_time[user]=process_time
        else:
            total_time[user]+=process_time
        if queue_start_time.get(user)==None:
            queue_start_time[user]=start_time
        queue_finish_time[user]=receive_time
        ch.basic_ack(delivery_tag=method.delivery_tag)
    channel.basic_consume(queue='fair_queue', on_message_callback=callback) 
    channel.start_consuming()

def get_key(val):
    for key, value in statistics_dic.items():
        if val == value:
            return key

def statistics_write():
    while True:
        inpt = input()
        if inpt == 'p':
            path = 'C:\\Users\\asus\\Desktop\\RabbitMQ 1'
            os.chdir(path)
            wb = load_workbook("statistics.xlsx")
            ws = wb.active
            print(first_message)
            print(statistics_dic)
            print(total_time)
            for i in queue_start_time:
                avarage.append((queue_finish_time.get(i)-queue_start_time.get(i))/statistics_dic.get(i))
            exit=False
            i=4
            count=0
            while exit==False:
                column = get_column_letter(i)
                if ws[column+str(4)].value==None and ws[column+str(5)].value==None and ws[column+str(6)].value==None and ws[column+str(7)].value==None and ws[column+str(8)].value==None and ws[column+str(9)].value==None and ws[column+str(10)].value==None and ws[column+str(11)].value==None and ws[column+str(12)].value==None and ws[column+str(13)].value==None:
                    for j in statistics_dic:
                        if get_key(statistics_dic[j])=='A':
                            line=4
                        if get_key(statistics_dic[j])=='B':
                            line=5
                        if get_key(statistics_dic[j])=='C':
                            line=6
                        if get_key(statistics_dic[j])=='D':
                            line=7
                        if get_key(statistics_dic[j])=='E':
                            line=8
                        if get_key(statistics_dic[j])=='F':
                            line=9
                        if get_key(statistics_dic[j])=='G':
                            line=10
                        if get_key(statistics_dic[j])=='H':
                            line=11
                        if get_key(statistics_dic[j])=='I':
                            line=12
                        if get_key(statistics_dic[j])=='M':
                            line=13
                        cell1=get_column_letter(i)+str(line)
                        cell2=get_column_letter(i+1)+str(line)
                        ws[cell1]=first_message[j]
                        ws[cell2]=avarage[count]
                        count+=1
                        exit=True
                else:
                    i+=6
            wb.save("statistics.xlsx")
            first_message.clear()
            statistics_dic.clear()
            total_time.clear()
            queue_start_time.clear()
            queue_finish_time.clear()
            avarage.clear()

t1 = threading.Thread(target=basic_consume)
t1.start()
t2 = threading.Thread(target=statistics_write)
t2.start()
