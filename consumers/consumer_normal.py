import os 
import pika
import time
import json
import requests    
import threading   
from openpyxl import Workbook,load_workbook

connection = pika.BlockingConnection(pika.ConnectionParameters(host='localhost'))
channel = connection.channel()
#os.chdir("..")

start_times = {}
receive_times = {}
statistics_dic = {}
queues = []
consumers = []
proc_time = 0.0

excel_receive_times = []
excel_start_times = []
excel_passing_times = []
excel_first_messages = {}

def queue_list(user='guest', password='guest', host='localhost', port=15672, virtual_host=None):
    url = 'http://%s:%s/api/queues/%s' % (host, port, virtual_host or '')
    response = requests.get(url, auth=(user, password))

def queue_list1(user='guest', password='guest', host='localhost', port=15672, virtual_host=None):
    path = 'C:\\Program Files\\RabbitMQ Server\\rabbitmq_server-3.8.14\sbin'
    os.chdir(path)
    a = os.popen("rabbitmqctl list_queues")
    writing = a.read()
    line = writing.splitlines()
    global queues
    for i in range (3, len(line)):
        queue_split = line[i].split("\t")
        queues.append(queue_split[0])

channel.basic_qos(prefetch_count=1)

def start_consume():
    while True:
        time.sleep(0.5)
        channel.start_consuming()

def basic_consume():
    while True:
        queue_list()
        def callback(ch, method, properties, body):
            receive_time = time.time()  
            received_data = json.loads(body)
            user = received_data['User']
            start_time = received_data['Time']
            excel_receive_times.append(receive_time)
            excel_start_times.append(start_time)
            excel_passing_times.append(receive_time-float(start_time))
            if excel_first_messages.get(user)==None:
                excel_first_messages[user]=receive_time-float(start_time)
            if statistics_dic.get(user)==None:
                statistics_dic[user] = 1
            else:
                statistics_dic[user] = statistics_dic.get(user) + 1
            if start_times.get(user)==None:
                receive_times[user] = receive_time
                start_times[user] = start_time
            ch.basic_ack(delivery_tag=method.delivery_tag)
        global consumers
        if not 'normal_queue' in consumers:
            channel.queue_declare(queue='normal_queue', durable=True)
            channel.basic_consume(queue='normal_queue', on_message_callback=callback, consumer_tag='normal_queue') 
            consumers.append('normal_queue')

def get_key(val):
    for key, value in excel_first_messages.items():
         if val == value:
             return key

def excel_statics():
    while True:
        inpt = input()
        if inpt == 'p':
            #path = 'C:\\Users\\asus\\Desktop\\RabbitMQ'
            #os.chdir(path)
            wb = load_workbook("statistics.xlsx")
            ws = wb.active
            toplam = 0
            ilk_mesaj = excel_receive_times[0]-float(excel_start_times[0])
            geçen_süre = excel_receive_times[len(excel_receive_times)-1]-float(excel_start_times[0])
            print(len(excel_receive_times))
            print(geçen_süre)
            for i in excel_passing_times:
                toplam  = toplam + i
            count = 1
            for row in ws:
                if row[1].value==None:
                    cell1 = 'B'+str(count)
                    cell3 = 'D'+str(count)
                    cell5 = 'F'+str(count)
                    cell6 = 'G'+str(count)
                    cell9 = 'J'+str(count)
                    cell11 = 'L'+str(count)
                    ws[cell1] = toplam/len(excel_receive_times)
                    ws[cell3] = ilk_mesaj
                    ws[cell5] = get_key(max(excel_first_messages.values()))
                    ws[cell6] = max(excel_first_messages.values())
                    ws[cell9] = len(excel_receive_times)
                    ws[cell11] = geçen_süre
                count = count + 1
            wb.save("statistics.xlsx")
            excel_first_messages.clear()
            excel_receive_times.clear()
            excel_start_times.clear()
            excel_passing_times.clear()

t1 = threading.Thread(target=basic_consume)
t1.start()
t2 = threading.Thread(target=start_consume)
t2.start()
t3 = threading.Thread(target=excel_statics)
t3.start()