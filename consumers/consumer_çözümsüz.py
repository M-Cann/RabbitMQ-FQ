import os 
import pika
import time
import json  
import threading   
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

connection = pika.BlockingConnection(pika.ConnectionParameters(host='localhost'))
channel = connection.channel()
channel.queue_declare(queue='normal_queue', durable=True)
channel.basic_qos(prefetch_count=1)

#os.chdir("..")
first_message = {}
statistics_dic = {}
all_process_total_time = {}
queue_start_time = {}
queue_finish_time = {}
avarage={} 
total_message_count = 0
producer_start_time = 0
process_finish_time = 0

def basic_consume():
    def callback(ch, method, properties, body):
        global message_count
        global producer_start_time
        global process_finish_time
        receive_time = time.time()  
        received_data = json.loads(body)
        user = received_data['User']
        start_time = received_data['Time']
        process_time=receive_time-start_time
        if producer_start_time==0:
            producer_start_time=start_time
        if statistics_dic.get(user)==None:
            statistics_dic[user]=1
        else:
            statistics_dic[user]=statistics_dic.get(user)+1
        if first_message.get(user)==None:
            first_message[user]=process_time
        if all_process_total_time.get(user)==None:
            all_process_total_time[user]=process_time
        else:
            all_process_total_time[user]+=process_time
        if queue_start_time.get(user)==None:
            queue_start_time[user]=start_time
        queue_finish_time[user]=receive_time
        process_finish_time = receive_time
        ch.basic_ack(delivery_tag=method.delivery_tag)
    channel.basic_consume(queue='normal_queue', on_message_callback=callback, consumer_tag='normal_queue') 
    channel.start_consuming()

def get_key(val):
    for key, value in statistics_dic.items():
        if val == value:
            return key

def statistics_write():
    while True:
        inpt = input()
        if inpt == 'p':
            #path = 'C:\\Users\\asus\\Desktop\\RabbitMQ 1'
            #os.chdir(path)
            wb = load_workbook("statistics.xlsx")
            ws = wb.active
            for i in queue_start_time:
                avarage[i]=all_process_total_time.get(i)/statistics_dic.get(i)
                #avarage[i]=(queue_finish_time.get(i)-queue_start_time.get(i))/statistics_dic.get(i)
            exit=False
            i=2        
            while exit==False:
                column = get_column_letter(i)
                if ws[column+str(4)].value==None and ws[column+str(5)].value==None and ws[column+str(6)].value==None and ws[column+str(7)].value==None and ws[column+str(8)].value==None and ws[column+str(9)].value==None and ws[column+str(10)].value==None and ws[column+str(11)].value==None and ws[column+str(12)].value==None and ws[column+str(13)].value==None:
                    for j in statistics_dic:
                        if get_key(statistics_dic[j])=='A':
                            line=4
                        if get_key(statistics_dic[j])=='B':
                            line=5
                        if get_key(statistics_dic[j])=='C':
                            line=6
                        if get_key(statistics_dic[j])=='D':
                            line=7
                        if get_key(statistics_dic[j])=='E':
                            line=8
                        if get_key(statistics_dic[j])=='F':
                            line=9
                        if get_key(statistics_dic[j])=='G':
                            line=10
                        if get_key(statistics_dic[j])=='H':
                            line=11
                        if get_key(statistics_dic[j])=='I':
                            line=12
                        if get_key(statistics_dic[j])=='M':
                            line=13
                        cell1=get_column_letter(i)+str(line)
                        cell2=get_column_letter(i+1)+str(line)
                        ws[cell1]=first_message[j]
                        ws[cell2]=avarage[j]
                        exit=True
                else:
                    i+=6

            wb2 = load_workbook("statistics2.xlsx")
            ws2 = wb2.active
            global total_message_count
            global process_finish_time
            global producer_start_time
            for i in statistics_dic:
                total_message_count += statistics_dic.get(i)
            exit=False
            i=2
            longest_first_message_time = max(first_message.values())
            longest_first_message_queue = [k for k, v in first_message.items() if v == longest_first_message_time]
            while exit==False:
                column = get_column_letter(i)
                if ws2[column+str(3)].value==None and ws2[column+str(4)].value==None and ws2[column+str(5)].value==None and ws2[column+str(6)].value==None and ws2[column+str(7)].value==None:
                    cell1=get_column_letter(i)+str(3)
                    cell2=get_column_letter(i)+str(4)
                    cell3=get_column_letter(i)+str(5)
                    cell4=get_column_letter(i)+str(6)
                    cell5=get_column_letter(i)+str(7)
                    ws2[cell1]=total_message_count
                    ws2[cell2]=process_finish_time-producer_start_time
                    ws2[cell3]=(process_finish_time-producer_start_time)/total_message_count
                    ws2[cell4]=longest_first_message_queue[0]
                    ws2[cell5]=longest_first_message_time
                    exit=True
                else:
                    i+=3
            print('total_message_count:', total_message_count)

            wb.save("statistics.xlsx")
            wb2.save("statistics2.xlsx")

            first_message.clear()
            statistics_dic.clear()
            all_process_total_time.clear()
            queue_start_time.clear()
            queue_finish_time.clear()
            avarage.clear()
            longest_first_message_queue.clear()
            longest_first_message_time = 0
            total_message_count = 0
            producer_start_time = 0
            process_finish_time = 0


t1 = threading.Thread(target=basic_consume)
t1.start()
t2 = threading.Thread(target=statistics_write)
t2.start()
